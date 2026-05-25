from __future__ import annotations

import csv
import io
import json
import zipfile
from dataclasses import asdict, dataclass, replace
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
from scipy import sparse
from scipy.integrate import solve_ivp
from scipy.io import loadmat

from .dynamics import get_model


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_ARTICLE_ZIP = Path("/Users/yangmingzhe/Downloads/NaturePhys2021-main.zip")
RESULT_SUBDIR = Path("results/network_revival_microbiome")
FIG_SUBDIR = Path("fig/network_revival_microbiome")


@dataclass(frozen=True)
class MicrobiomeParameters:
    positive_weight: float = 30.0
    competition_weight: float = 1.0
    eco_f: float = 5.0
    eco_b: float = 3.0
    eco_c: float = 3.0
    eco_k: float = 10.0
    upper_free_value: float = 10.0
    collapsed_free_value: float = 0.1
    active_threshold: float = 2.0
    delta: float = 10.0
    success_threshold: float = 5.0
    reach_threshold: float = 1.122
    dt: float = 0.05
    t_max: float = 20.0
    steady_tol: float = 1.0e-3


@dataclass(frozen=True)
class MicrobiomeNetworks:
    competition: np.ndarray
    complementarity: np.ndarray
    norm_import: np.ndarray
    source: str


@dataclass(frozen=True)
class ActiveMicrobiomeNetwork:
    active_adjacency: np.ndarray
    active_indices: np.ndarray
    active_upper_state: np.ndarray


def load_microbiome_networks(article_zip: str | Path = DEFAULT_ARTICLE_ZIP) -> MicrobiomeNetworks:
    """Load numeric microbiome matrices from the article archive."""

    zip_path = Path(article_zip).expanduser()
    if not zip_path.exists():
        raise FileNotFoundError(f"Article zip not found: {zip_path}")

    with zipfile.ZipFile(zip_path) as zf:
        raw = zf.read("NaturePhys2021-main/data/MicrobiomeNetworks.mat")
    data = loadmat(
        io.BytesIO(raw),
        variable_names=["competition", "complementarity", "norm_import"],
    )
    return MicrobiomeNetworks(
        competition=np.asarray(data["competition"], dtype=float),
        complementarity=np.asarray(data["complementarity"], dtype=float),
        norm_import=np.asarray(data["norm_import"], dtype=float),
        source=str(zip_path),
    )


def load_reference_ranked_species(article_zip: str | Path = DEFAULT_ARTICLE_ZIP) -> list[str]:
    """Read species labels from the paper-generated ranked Excel table."""

    from openpyxl import load_workbook

    zip_path = Path(article_zip).expanduser()
    with zipfile.ZipFile(zip_path) as zf:
        raw = zf.read("NaturePhys2021-main/output/Figure6/i.xlsx")
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    return [str(row[1]) for row in rows if row and row[1] is not None]


def build_signed_microbiome_adjacency(
    data: MicrobiomeNetworks,
    params: MicrobiomeParameters = MicrobiomeParameters(),
) -> np.ndarray:
    """Return A = wp * complementarity - wc * competition."""

    adjacency = params.positive_weight * data.complementarity - params.competition_weight * data.competition
    np.fill_diagonal(adjacency, 0.0)
    return adjacency.astype(float, copy=False)


def filter_active_nodes(
    adjacency: np.ndarray,
    upper_state: np.ndarray,
    *,
    threshold: float,
) -> ActiveMicrobiomeNetwork:
    """Keep nodes whose upper-state activity is above the paper threshold."""

    state = np.asarray(upper_state, dtype=float).reshape(-1)
    active_indices = np.flatnonzero(state > float(threshold))
    active_adjacency = np.asarray(adjacency, dtype=float)[np.ix_(active_indices, active_indices)]
    return ActiveMicrobiomeNetwork(
        active_adjacency=active_adjacency,
        active_indices=active_indices.astype(int, copy=False),
        active_upper_state=state[active_indices],
    )


def solve_ecological_steady_state(
    adjacency,
    params: MicrobiomeParameters = MicrobiomeParameters(),
    *,
    free_value: float,
    fixed_node: int | None = None,
    fixed_value: float | None = None,
) -> np.ndarray:
    """Integrate the ecological dynamics to a steady state or t_max."""

    model = get_model("Eco", F=params.eco_f, B=params.eco_b, C=params.eco_c, K=params.eco_k)
    matrix = sparse.csr_matrix(adjacency)
    n_nodes = int(matrix.shape[0])
    x = np.full(n_nodes, float(free_value), dtype=float)
    fixed_mask = np.zeros(n_nodes, dtype=bool)
    if fixed_node is not None:
        fixed = int(fixed_node)
        fixed_mask[fixed] = True
        x[fixed] = float(fixed_value)

    def rhs(_, state):
        dx = model["M0"](state) + model["M1"](state) * matrix.dot(model["M2"](state))
        if fixed_node is not None:
            dx[fixed_mask] = 0.0
        return dx

    solution = solve_ivp(
        rhs,
        (0.0, float(params.t_max)),
        x,
        method="RK45",
        rtol=1.0e-3,
        atol=1.0e-6,
        max_step=float(params.dt),
    )
    if not solution.success:
        raise RuntimeError(f"Ecological ODE integration failed: {solution.message}")
    x_ss = np.asarray(solution.y[:, -1], dtype=float)
    if fixed_node is not None:
        x_ss[fixed_mask] = float(fixed_value)
    return x_ss


def compute_node_ignition_states(
    adjacency: np.ndarray,
    params: MicrobiomeParameters = MicrobiomeParameters(),
    *,
    node_indices: Iterable[int] | None = None,
) -> dict[int, float]:
    """Force each requested active-network node and return final mean state."""

    nodes = list(range(adjacency.shape[0])) if node_indices is None else [int(node) for node in node_indices]
    states: dict[int, float] = {}
    for node in nodes:
        x_ss = solve_ecological_steady_state(
            adjacency,
            params,
            free_value=params.collapsed_free_value,
            fixed_node=node,
            fixed_value=params.delta,
        )
        states[node] = float(np.mean(x_ss))
    return states


def rank_nodes_by_positive_reach(
    adjacency: np.ndarray,
    *,
    active_indices: np.ndarray,
    states: np.ndarray,
    threshold: float,
    species_names: dict[int, str] | None = None,
    success_threshold: float = 5.0,
) -> list[dict[str, object]]:
    """Rank nodes by reachable size in the thresholded positive digraph A'."""

    reach_graph = np.asarray(adjacency, dtype=float).T >= float(threshold)
    n_nodes = reach_graph.shape[0]
    active_indices = np.asarray(active_indices, dtype=int)
    state_values = np.asarray(states, dtype=float)
    names = species_names or {}

    rows: list[dict[str, object]] = []
    for source in range(n_nodes):
        reached = _reachable_count(reach_graph, source)
        species_index = int(active_indices[source])
        state = float(state_values[source])
        rows.append(
            {
                "active_node": int(source),
                "species_index": species_index,
                "species_name": names.get(species_index, f"species_{species_index + 1}"),
                "tree_size": int(reached),
                "state": state,
                "success": bool(state > float(success_threshold)),
            }
        )

    rows.sort(key=lambda row: (-int(row["tree_size"]), int(row["active_node"])))
    for rank, row in enumerate(rows, start=1):
        row["rank"] = rank
    return rows


def run_microbiome_reproduction(
    *,
    article_zip: str | Path = DEFAULT_ARTICLE_ZIP,
    output_dir: str | Path = ".",
    params: MicrobiomeParameters = MicrobiomeParameters(),
    max_nodes: int | None = None,
    node_indices: Iterable[int] | None = None,
    force: bool = False,
) -> dict[str, object]:
    """Run the microbiome point-ignition reproduction and write artifacts."""

    root = Path(output_dir)
    results_dir = root / RESULT_SUBDIR
    fig_dir = root / FIG_SUBDIR
    paths = _artifact_paths(results_dir, fig_dir)
    if not force and all(path.exists() for path in paths.values()):
        return {"results_dir": results_dir, "fig_dir": fig_dir, "metadata": _read_json(paths["metadata"])}

    results_dir.mkdir(parents=True, exist_ok=True)
    fig_dir.mkdir(parents=True, exist_ok=True)

    data = load_microbiome_networks(article_zip)
    full_adjacency = build_signed_microbiome_adjacency(data, params)
    upper_state = solve_ecological_steady_state(full_adjacency, params, free_value=params.upper_free_value)
    active = filter_active_nodes(full_adjacency, upper_state, threshold=params.active_threshold)

    if node_indices is None:
        selected_nodes = list(range(active.active_adjacency.shape[0]))
    else:
        selected_nodes = [int(node) for node in node_indices]
    if max_nodes is not None:
        selected_nodes = selected_nodes[: int(max_nodes)]

    state_by_node = compute_node_ignition_states(active.active_adjacency, params, node_indices=selected_nodes)
    states = np.full(active.active_adjacency.shape[0], np.nan, dtype=float)
    for node, state in state_by_node.items():
        states[int(node)] = float(state)

    species_names: dict[int, str] = {}
    ranked_all = rank_nodes_by_positive_reach(
        active.active_adjacency,
        active_indices=active.active_indices,
        states=np.nan_to_num(states, nan=0.0),
        threshold=params.reach_threshold,
        species_names=species_names,
        success_threshold=params.success_threshold,
    )
    _attach_reference_rank_names(ranked_all, article_zip)
    ranked_selected = [row for row in ranked_all if int(row["active_node"]) in set(selected_nodes)]

    np.savez_compressed(
        paths["active_network"],
        full_adjacency=full_adjacency,
        upper_state=upper_state,
        active_adjacency=active.active_adjacency,
        active_indices=active.active_indices,
        active_upper_state=active.active_upper_state,
        evaluated_active_nodes=np.asarray(selected_nodes, dtype=int),
        ignition_states=states,
    )
    _write_ignition_states_csv(paths["states"], active.active_indices, states, species_names, selected_nodes)
    _write_ranked_csv(paths["ranked"], ranked_selected)
    metadata = {
        "experiment": "network_revival_microbiome_point_ignition",
        "article_zip": str(Path(article_zip).expanduser()),
        "parameters": asdict(params),
        "full_node_count": int(full_adjacency.shape[0]),
        "active_node_count": int(active.active_adjacency.shape[0]),
        "evaluated_node_count": int(len(selected_nodes)),
        "release_after_forcing": False,
        "reference_table": "NaturePhys2021-main/output/Figure6/i.xlsx",
    }
    paths["metadata"].write_text(json.dumps(metadata, indent=2), encoding="utf-8")

    plot_signed_adjacency(full_adjacency, paths["full_heatmap"], max_size=300)
    plot_signed_adjacency(active.active_adjacency, paths["active_heatmap"], max_size=300)
    plot_ranked_ignition_success(ranked_selected, paths["ranked_success"])
    plot_tree_size_vs_state(ranked_selected, paths["tree_state"])

    return {
        "results_dir": results_dir,
        "fig_dir": fig_dir,
        "metadata": metadata,
        "ranked_rows": ranked_selected,
        "active": active,
        "states": states,
    }


def plot_signed_adjacency(adjacency: np.ndarray, output_path: Path, *, max_size: int | None = None) -> None:
    matrix = np.asarray(adjacency, dtype=float)
    if max_size is not None:
        matrix = matrix[:max_size, :max_size]
    signs = np.sign(matrix)
    fig, ax = plt.subplots(figsize=(5.0, 4.8), constrained_layout=True)
    ax.imshow(signs, cmap=plt.matplotlib.colors.ListedColormap(["#cc4c27", "#ffffff", "#0072b2"]), vmin=-1, vmax=1)
    ax.set_xticks([])
    ax.set_yticks([])
    for spine in ax.spines.values():
        spine.set_visible(False)
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def plot_ranked_ignition_success(rows: list[dict[str, object]], output_path: Path) -> None:
    ranks = np.array([int(row["rank"]) for row in rows], dtype=int)
    success = np.array([1.0 if bool(row["success"]) else 0.0 for row in rows], dtype=float)
    fig, ax = plt.subplots(figsize=(7.0, 3.2), constrained_layout=True)
    if ranks.size:
        ax.scatter(ranks, success, s=18, color="#0072b2", edgecolors="none")
    ax.set_xlabel("Rank")
    ax.set_ylabel("Ignition success")
    ax.set_ylim(-0.1, 1.1)
    ax.set_yticks([0, 1])
    ax.grid(axis="y", color="0.88", linewidth=0.8)
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def plot_tree_size_vs_state(rows: list[dict[str, object]], output_path: Path) -> None:
    tree_size = np.array([int(row["tree_size"]) for row in rows], dtype=float)
    state = np.array([float(row["state"]) for row in rows], dtype=float)
    fig, ax = plt.subplots(figsize=(5.2, 3.6), constrained_layout=True)
    if tree_size.size:
        ax.scatter(tree_size, state, s=22, color="#009e73", edgecolors="none")
    ax.axhline(5.0, color="0.35", linewidth=1.0, linestyle="--")
    ax.set_xlabel("Tree size")
    ax.set_ylabel("Final mean state")
    ax.grid(color="0.9", linewidth=0.8)
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _reachable_count(graph: np.ndarray, source: int) -> int:
    seen = np.zeros(graph.shape[0], dtype=bool)
    stack = [int(source)]
    seen[int(source)] = True
    while stack:
        node = stack.pop()
        for neighbor in np.flatnonzero(graph[node]):
            if not seen[neighbor]:
                seen[neighbor] = True
                stack.append(int(neighbor))
    return int(seen.sum())


def _attach_reference_rank_names(rows: list[dict[str, object]], article_zip: str | Path) -> None:
    try:
        ranked_names = load_reference_ranked_species(article_zip)
    except Exception:
        return
    if len(ranked_names) != len(rows):
        return
    for row, name in zip(rows, ranked_names):
        row["species_name"] = name


def _artifact_paths(results_dir: Path, fig_dir: Path) -> dict[str, Path]:
    return {
        "active_network": results_dir / "active_network.npz",
        "states": results_dir / "node_ignition_states.csv",
        "ranked": results_dir / "node_ignition_ranked.csv",
        "metadata": results_dir / "metadata.json",
        "full_heatmap": fig_dir / "signed_full_adjacency.png",
        "active_heatmap": fig_dir / "signed_active_adjacency.png",
        "ranked_success": fig_dir / "ranked_ignition_success.png",
        "tree_state": fig_dir / "tree_size_vs_state.png",
    }


def _write_ignition_states_csv(
    path: Path,
    active_indices: np.ndarray,
    states: np.ndarray,
    species_names: dict[int, str],
    selected_nodes: Iterable[int],
) -> None:
    selected = set(int(node) for node in selected_nodes)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["active_node", "species_index", "species_name", "state", "evaluated"],
        )
        writer.writeheader()
        for active_node, species_index in enumerate(active_indices):
            writer.writerow(
                {
                    "active_node": int(active_node),
                    "species_index": int(species_index),
                    "species_name": species_names.get(int(species_index), f"species_{int(species_index) + 1}"),
                    "state": "" if np.isnan(states[active_node]) else float(states[active_node]),
                    "evaluated": int(active_node) in selected,
                }
            )


def _write_ranked_csv(path: Path, rows: list[dict[str, object]]) -> None:
    fieldnames = ["rank", "active_node", "species_index", "species_name", "tree_size", "state", "success"]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row[key] for key in fieldnames})


def _read_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))
